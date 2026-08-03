import io
import datetime
from io import BytesIO

from sqlalchemy.orm import Session
from fastapi import HTTPException, status
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.repositories.project_repository import ProjectRepository
from app.repositories.workspace_repository import WorkspaceRepository


class ReportService:
    def __init__(self, db: Session):
        self._db = db
        self._project_repo = ProjectRepository(db)
        self._workspace_repo = WorkspaceRepository(db)

    def _get_project_data(self, workspace_id: uuid.UUID, project_id: uuid.UUID):
        project = self._project_repo.get_by_id(project_id)
        if not project or project.workspace_id != workspace_id:
            raise HTTPException(status_code=404, detail="Project not found")

        workspace = self._workspace_repo.get_by_id(workspace_id)
        
        meta = project.metadata_ or {}
        comparison = meta.get("comparison", {})
        compliance = meta.get("compliance", {})
        recommendation = meta.get("recommendation", {})
        human_review = meta.get("human_review", "Pending Human Approval")

        # Fallbacks
        vendors = comparison.get("vendor_rankings", [])
        quotations = project.quotations

        return {
            "project": project,
            "workspace": workspace,
            "comparison": comparison,
            "compliance": compliance,
            "recommendation": recommendation,
            "human_review": human_review,
            "vendors": vendors,
            "quotations": quotations
        }

    def generate_excel_report(self, workspace_id: uuid.UUID, project_id: uuid.UUID) -> BytesIO:
        data = self._get_project_data(workspace_id, project_id)
        wb = Workbook()

        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        self._write_excel_sheet1(ws1, data)

        # Sheet 2: Vendor Comparison
        ws2 = wb.create_sheet(title="Vendor Comparison")
        self._write_excel_sheet2(ws2, data)

        # Sheet 3: Compliance Audit
        ws3 = wb.create_sheet(title="Compliance Audit")
        self._write_excel_sheet3(ws3, data)

        # Sheet 4: Recommendation
        ws4 = wb.create_sheet(title="Recommendation")
        self._write_excel_sheet4(ws4, data)

        # Sheet 5: Quotation Details
        ws5 = wb.create_sheet(title="Quotation Details")
        self._write_excel_sheet5(ws5, data)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_excel_sheet1(self, ws, data):
        ws.append(["EXECUTIVE SUMMARY"])
        ws["A1"].font = Font(bold=True, size=14)
        
        project = data["project"]
        rec = data["recommendation"]
        vendors = data["vendors"]
        comp = data["compliance"].get("quotation_results", [])
        
        compliant_count = sum(1 for c in comp if c.get("status") == "COMPLIANT")
        non_compliant_count = sum(1 for c in comp if c.get("status") == "NON_COMPLIANT")

        ws.append(["Project Name", project.name])
        ws.append(["Total Vendors", len(vendors)])
        ws.append(["Total Quotations", len(data["quotations"])])
        ws.append(["Compliant Vendors", compliant_count])
        ws.append(["Non-Compliant Vendors", non_compliant_count])
        ws.append(["Recommended Vendor", rec.get("recommended_vendor", "None")])
        ws.append(["Confidence Score", rec.get("confidence_score", 0)])
        ws.append(["Human Review Status", data["human_review"]])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 25

    def _write_excel_sheet2(self, ws, data):
        headers = ["Vendor", "Grand Total", "Discount", "Delivery (Days)", "Warranty", "Overall Rank"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for v in data["vendors"]:
            ws.append([
                v.get("vendor_name", ""),
                v.get("grand_total", 0),
                v.get("discount", ""),
                v.get("delivery_time", ""),
                v.get("warranty", ""),
                v.get("rank", "")
            ])

    def _write_excel_sheet3(self, ws, data):
        headers = ["Vendor Name", "Compliance Status", "Errors", "Warnings", "Issues Breakdown"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for qr in data["compliance"].get("quotation_results", []):
            issues = qr.get("issues", [])
            issue_text = ", ".join([f"{i.get('check_name')}" for i in issues])
            ws.append([
                qr.get("vendor_name", ""),
                qr.get("status", ""),
                qr.get("failed_checks", 0),
                qr.get("warning_count", 0),
                issue_text
            ])

    def _write_excel_sheet4(self, ws, data):
        ws.append(["RECOMMENDATION"])
        ws["A1"].font = Font(bold=True)
        rec = data["recommendation"]
        ws.append(["Recommended Vendor", rec.get("recommended_vendor", "None")])
        ws.append(["Confidence", rec.get("confidence_score", 0)])
        ws.append(["Reasoning", rec.get("reasoning", "")])
        
        ws.append([])
        ws.append(["Strengths"])
        for s in rec.get("strengths", []):
            ws.append(["-", s])
            
        ws.append([])
        ws.append(["Risks"])
        for r in rec.get("risks", []):
            ws.append(["-", r])

    def _write_excel_sheet5(self, ws, data):
        headers = ["Vendor", "File", "Status", "Created"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        for q in data["quotations"]:
            ws.append([
                q.vendor_name or "Unknown",
                q.file_name,
                q.status.value,
                str(q.created_at)
            ])


    def generate_pdf_report(self, workspace_id: uuid.UUID, project_id: uuid.UUID) -> BytesIO:
        data = self._get_project_data(workspace_id, project_id)
        
        output = BytesIO()
        doc = SimpleDocTemplate(
            output, 
            pagesize=letter,
            rightMargin=72, leftMargin=72,
            topMargin=72, bottomMargin=72
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=24, spaceAfter=20, alignment=1)
        header_style = ParagraphStyle('HeaderStyle', parent=styles['Heading2'], fontSize=16, spaceAfter=15, textColor=colors.darkblue)
        normal_style = styles['Normal']
        
        story = []

        # PAGE 1: COVER PAGE
        story.append(Spacer(1, 100))
        story.append(Paragraph("ProcureAI", ParagraphStyle('PLogo', fontSize=30, textColor=colors.HexColor("#2563eb"), alignment=1, spaceAfter=40)))
        story.append(Paragraph("Executive Procurement Analysis Report", title_style))
        story.append(Spacer(1, 40))
        story.append(Paragraph(f"<b>Project:</b> {data['project'].name}", ParagraphStyle('Cen', alignment=1, fontSize=14)))
        story.append(Paragraph(f"<b>Workspace:</b> {data['workspace'].name}", ParagraphStyle('Cen', alignment=1, fontSize=12, spaceBefore=10)))
        story.append(Spacer(1, 100))
        story.append(Paragraph(f"Generated Date: {datetime.datetime.now().strftime('%Y-%m-%d')}", ParagraphStyle('Cen', alignment=1)))
        story.append(Paragraph(f"Generated Time: {datetime.datetime.now().strftime('%H:%M:%S')}", ParagraphStyle('Cen', alignment=1)))
        story.append(Paragraph(f"Report ID: {str(uuid.uuid4())[:8]}", ParagraphStyle('Cen', alignment=1)))
        story.append(PageBreak())

        # PAGE 2: EXECUTIVE SUMMARY
        story.append(Paragraph("Executive Summary", header_style))
        exec_data = [
            ["Project Name", data["project"].name],
            ["Total Vendors", str(len(data["vendors"]))],
            ["Total Quotations", str(len(data["quotations"]))],
            ["Recommended Vendor", data["recommendation"].get("recommended_vendor", "None")],
            ["Confidence Score", f"{data['recommendation'].get('confidence_score', 0)}%"],
            ["Human Review Status", data["human_review"]]
        ]
        t = Table(exec_data, colWidths=[200, 250])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
            ('BOX', (0,0), (-1,-1), 0.25, colors.black),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(t)
        story.append(PageBreak())

        # PAGE 3: AI EXECUTIVE REASONING
        story.append(Paragraph("AI Executive Reasoning", header_style))
        reasoning = data["recommendation"].get("reasoning", "No algorithmic reasoning generated.")
        story.append(Paragraph(reasoning, normal_style))
        story.append(PageBreak())

        # PAGE 4: DECISION MATRIX
        story.append(Paragraph("Decision Matrix", header_style))
        matrix_data = [["Vendor", "Grand Total", "Delivery", "Warranty", "Rank"]]
        for v in data["vendors"]:
            matrix_data.append([
                v.get("vendor_name", ""),
                str(v.get("grand_total", "")),
                str(v.get("delivery_time", "")),
                str(v.get("warranty", "")),
                str(v.get("rank", ""))
            ])
        t_matrix = Table(matrix_data)
        t_matrix.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2563eb")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
            ('BOX', (0,0), (-1,-1), 0.25, colors.black),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(t_matrix)
        story.append(PageBreak())

        # PAGE 5: COMPLIANCE AUDIT
        story.append(Paragraph("Compliance Audit", header_style))
        for qr in data["compliance"].get("quotation_results", []):
            story.append(Paragraph(f"<b>{qr.get('vendor_name', 'Unknown')}</b> - {qr.get('status', '')}", ParagraphStyle('H3', fontSize=12, spaceAfter=8)))
            issues = qr.get("issues", [])
            if not issues:
                story.append(Paragraph("No compliance issues detected.", normal_style))
            for issue in issues:
                story.append(Paragraph(f"• {issue.get('check_name')}: {issue.get('message')}", normal_style))
            story.append(Spacer(1, 15))
        story.append(PageBreak())

        # PAGE 6: RECOMMENDATION
        story.append(Paragraph("Recommendation Profile", header_style))
        rec = data["recommendation"]
        story.append(Paragraph(f"<b>Top Selection: {rec.get('recommended_vendor', 'None')}</b>", normal_style))
        story.append(Spacer(1, 10))
        story.append(Paragraph("Strengths:", ParagraphStyle('H3', fontSize=12, spaceAfter=8)))
        for s in rec.get("strengths", []):
            story.append(Paragraph(f"• {s}", normal_style))
        
        story.append(Spacer(1, 10))
        story.append(Paragraph("Associated Risks:", ParagraphStyle('H3', fontSize=12, spaceAfter=8)))
        for r in rec.get("risks", []):
            story.append(Paragraph(f"• {r}", normal_style))
        story.append(PageBreak())

        # PAGE 7: WHY OTHER VENDORS WERE NOT SELECTED
        story.append(Paragraph("Why Other Vendors Were Not Selected", header_style))
        alt_names = rec.get("alternatives", [])
        if alt_names:
            for alt in alt_names:
                story.append(Paragraph(f"<b>{alt}</b>", ParagraphStyle('H3', fontSize=12, spaceAfter=8, spaceBefore=10)))
                story.append(Paragraph("Reasons Not Selected: Did not meet the overall comparative threshold versus the optimal selection during cost, compliance, and negotiation analyses.", normal_style))
        else:
            story.append(Paragraph("No viable alternatives were ranked.", normal_style))
        story.append(PageBreak())

        # PAGE 8: HUMAN REVIEW
        story.append(Paragraph("Human Review", header_style))
        story.append(Paragraph(f"Status: {data['human_review']}", normal_style))
        story.append(PageBreak())

        # PAGE 9: APPENDIX
        story.append(Paragraph("Appendix (Quotation Audit Trail)", header_style))
        for q in data["quotations"]:
            story.append(Paragraph(f"<b>Vendor:</b> {q.vendor_name}", normal_style))
            story.append(Paragraph(f"<b>File:</b> {q.file_name}", normal_style))
            story.append(Paragraph(f"<b>Status:</b> {q.status.value}", normal_style))
            story.append(Spacer(1, 15))
        story.append(PageBreak())

        # FINAL PAGE
        story.append(Spacer(1, 100))
        story.append(Paragraph("<b>AI Recommendation Disclaimer</b>", normal_style))
        story.append(Paragraph("This report was generated by ProcureAI using extracted quotation data, deterministic comparison algorithms, compliance validation, recommendation logic, and human review inputs.", normal_style))
        story.append(Paragraph("The recommendation serves as a decision-support tool and should be reviewed by authorized procurement personnel before final approval.", normal_style))

        doc.build(story)
        output.seek(0)
        return output
